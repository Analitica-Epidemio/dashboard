"""
Preview endpoint - upload and analyze file without processing (OPTIMIZED)
"""

import logging
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any, List

import pandas as pd
from fastapi import Depends, File, HTTPException, UploadFile, status
from pydantic import BaseModel

from app.core.schemas.response import SuccessResponse
from app.core.security import RequireAnyRole
from app.domains.autenticacion.models import User
from app.features.procesamiento_archivos.config.columns import REQUIRED_COLUMNS

logger = logging.getLogger(__name__)

# Directorio temporal para archivos en preview
TEMP_UPLOAD_DIR = Path(tempfile.gettempdir()) / "epidemio_uploads"
TEMP_UPLOAD_DIR.mkdir(exist_ok=True)


class SheetPreviewData(BaseModel):
    """Preview data for a single sheet."""
    name: str
    columns: List[str]
    row_count: int
    preview_rows: List[List[Any]]  # Any para permitir tipos mixtos
    is_valid: bool
    missing_columns: List[str]


class FilePreviewResponse(BaseModel):
    """Response with file preview data."""
    upload_id: str
    filename: str
    file_size: int
    sheets: List[SheetPreviewData]
    valid_sheets_count: int
    total_sheets_count: int


def validate_columns(columns: List[str]) -> bool:
    """Check if sheet has all required columns."""
    columns_upper = {col.upper().strip() for col in columns}
    required_upper = {col.upper() for col in REQUIRED_COLUMNS}
    return required_upper.issubset(columns_upper)


def get_missing_columns(columns: List[str]) -> List[str]:
    """Get list of missing required columns."""
    columns_upper = {col.upper().strip() for col in columns}
    required_upper = {col.upper(): col for col in REQUIRED_COLUMNS}
    missing = [required_upper[req] for req in required_upper.keys() if req not in columns_upper]
    return missing


def clean_preview_data(df: pd.DataFrame, max_rows: int = 10) -> List[List[Any]]:
    """
    Convert DataFrame to serializable preview data.

    OPTIMIZED: Convert all values to Python native types (str, int, float, None)
    to avoid pandas/numpy serialization issues.
    """
    # Take only first N rows
    df_preview = df.head(max_rows)

    # Convert to list and clean values
    rows = []
    for _, row in df_preview.iterrows():
        clean_row = []
        for val in row:
            if pd.isna(val):
                clean_row.append("")
            elif isinstance(val, (int, bool)):
                clean_row.append(int(val))
            elif isinstance(val, float):
                clean_row.append(float(val))
            else:
                clean_row.append(str(val))
        rows.append(clean_row)

    return rows


def get_total_row_count(file_path: Path, sheet_name: str = None) -> int:
    """
    Get total row count efficiently without loading all data.

    For CSV: Uses chunk reading
    For Excel: Uses openpyxl to get max_row
    """
    file_ext = file_path.suffix.lower()

    if file_ext == '.csv':
        # Count lines efficiently
        logger.debug("🔢 Counting CSV rows using chunks...")
        chunk_start = time.time()
        row_count = 0
        for chunk in pd.read_csv(file_path, chunksize=10000):
            row_count += len(chunk)
        chunk_duration = time.time() - chunk_start
        logger.debug(f"🔢 CSV chunk counting took {chunk_duration:.2f}s")
        return row_count
    else:
        # Excel - use openpyxl for efficient counting
        logger.debug(f"🔢 Counting Excel rows using openpyxl for sheet '{sheet_name}'...")
        openpyxl_start = time.time()
        from openpyxl import load_workbook
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Handle empty sheets or sheets with only header
        if ws.max_row is None or ws.max_row <= 1:
            row_count = 0
        else:
            row_count = ws.max_row - 1  # -1 to exclude header

        wb.close()
        openpyxl_duration = time.time() - openpyxl_start
        logger.debug(f"🔢 openpyxl counting took {openpyxl_duration:.2f}s")
        return row_count


async def preview_uploaded_file(
    file: UploadFile = File(..., description="Archivo Excel o CSV epidemiológico"),
    current_user: User = Depends(RequireAnyRole())
):
    """
    Preview uploaded file - OPTIMIZED VERSION.

    **Optimizations:**
    - Read each sheet ONLY ONCE with nrows limit
    - No full file scanning for row counts (too slow)
    - Convert all data to native Python types immediately
    - Minimal memory footprint

    **Returns:** Upload ID + sheet previews
    """

    logger.info(f"📤 Preview request - filename: {file.filename}, user: {current_user.email}")

    # Validate file type
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Nombre de archivo no válido"
        )

    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ['.xlsx', '.xls', '.csv']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Formato no soportado: {file_ext}. Use Excel (.xlsx, .xls) o CSV (.csv)"
        )

    try:
        start_time = time.time()

        # Generate unique upload ID
        upload_id = str(uuid.uuid4())

        # Save to temp file
        temp_file_path = TEMP_UPLOAD_DIR / f"{upload_id}_{file.filename}"

        logger.info(f"💾 Saving to temp: {temp_file_path}")
        save_start = time.time()

        with open(temp_file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        file_size = len(content)
        save_duration = time.time() - save_start
        logger.info(f"✅ File saved - size: {file_size / (1024*1024):.2f} MB - took {save_duration:.2f}s")

        # Analyze file structure
        sheets_data = []

        if file_ext == '.csv':
            # CSV file - single "sheet"
            logger.info("📊 Analyzing CSV file")
            csv_start = time.time()

            # Read first 10 rows for preview
            read_start = time.time()
            df = pd.read_csv(temp_file_path, nrows=10)
            read_duration = time.time() - read_start
            logger.info(f"📖 Read preview rows - took {read_duration:.2f}s")

            columns = df.columns.tolist()
            preview_rows = clean_preview_data(df, max_rows=10)

            # Get total row count efficiently
            logger.info("🔢 Counting total rows...")
            count_start = time.time()
            total_rows = get_total_row_count(temp_file_path)
            count_duration = time.time() - count_start
            logger.info(f"✅ Total rows: {total_rows:,} - took {count_duration:.2f}s")

            csv_duration = time.time() - csv_start
            logger.info(f"✅ CSV analysis complete - took {csv_duration:.2f}s")

            is_valid = validate_columns(columns)
            missing = get_missing_columns(columns) if not is_valid else []

            sheets_data.append(SheetPreviewData(
                name=Path(file.filename).stem,
                columns=columns,
                row_count=total_rows,
                preview_rows=preview_rows,
                is_valid=is_valid,
                missing_columns=missing
            ))

        else:
            # Excel file - multiple sheets
            logger.info("📊 Analyzing Excel file")
            excel_start = time.time()

            # OPTIMIZATION: Open Excel file ONCE and reuse
            # Try to use calamine (faster Rust-based engine), fallback to openpyxl
            logger.info("📂 Opening Excel file (single open)...")
            open_start = time.time()

            try:
                # Try calamine first (3-5x faster than openpyxl)
                excel_file = pd.ExcelFile(temp_file_path, engine='calamine')
                logger.info("⚡ Using calamine engine (fast)")
            except Exception:
                # Fallback to openpyxl
                excel_file = pd.ExcelFile(temp_file_path, engine='openpyxl')
                logger.info("🐢 Using openpyxl engine (slower)")

            open_duration = time.time() - open_start
            logger.info(f"✅ File opened - took {open_duration:.2f}s")

            sheet_names = excel_file.sheet_names
            logger.info(f"Found {len(sheet_names)} sheets: {sheet_names}")

            try:
                # Process each sheet using the SAME open file handle
                for sheet_name in sheet_names:
                    sheet_start = time.time()
                    logger.info(f"📄 Processing sheet: {sheet_name}")

                    # Read ALL data for accurate count (calamine is fast enough)
                    read_start = time.time()

                    df_full = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name
                    )

                    read_duration = time.time() - read_start
                    total_rows = len(df_full)
                    logger.info(f"📖 Read {total_rows:,} rows - took {read_duration:.2f}s")

                    columns = df_full.columns.tolist()
                    preview_rows = clean_preview_data(df_full, max_rows=10)

                    sheet_duration = time.time() - sheet_start
                    logger.info(f"✅ Sheet '{sheet_name}': {total_rows:,} rows - took {sheet_duration:.2f}s")

                    is_valid = validate_columns(columns)
                    missing = get_missing_columns(columns) if not is_valid else []

                    sheets_data.append(SheetPreviewData(
                        name=sheet_name,
                        columns=columns,
                        row_count=total_rows,
                        preview_rows=preview_rows,
                        is_valid=is_valid,
                        missing_columns=missing
                    ))
            finally:
                excel_file.close()

            excel_duration = time.time() - excel_start
            logger.info(f"✅ Excel analysis complete - took {excel_duration:.2f}s")

        # Build response
        valid_count = sum(1 for s in sheets_data if s.is_valid)
        total_count = len(sheets_data)

        total_duration = time.time() - start_time
        logger.info(f"✅ Preview complete - {valid_count}/{total_count} valid sheets - TOTAL TIME: {total_duration:.2f}s")

        if valid_count == 0:
            logger.warning("⚠️ No valid sheets found")

        response = FilePreviewResponse(
            upload_id=upload_id,
            filename=file.filename,
            file_size=file_size,
            sheets=sheets_data,
            valid_sheets_count=valid_count,
            total_sheets_count=total_count
        )

        return SuccessResponse(data=response)

    except pd.errors.EmptyDataError:
        logger.error("❌ Empty file")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo está vacío"
        )

    except Exception as e:
        logger.error(f"❌ Error analyzing file: {str(e)}", exc_info=True)

        # Clean up temp file on error
        if 'temp_file_path' in locals() and temp_file_path.exists():
            temp_file_path.unlink()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error analizando archivo: {str(e)}"
        )
